from pathlib import Path
from decimal import Decimal

import pytest
from openpyxl import Workbook

from kz_tax_report.annual_rates import AnnualRateError, AnnualRateProvider


def make_workbook(path: Path, *, rate: object = 521.59) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2025"
    sheet.cell(3, 3, "Код")
    sheet.cell(3, 4, "Январь")
    sheet.cell(3, 5, 2025)
    sheet.cell(4, 2, "Доллар США")
    sheet.cell(4, 3, "USD")
    sheet.cell(4, 5, rate)
    workbook.save(path)


def test_annual_rate_provider_reads_year_and_currency(tmp_path: Path) -> None:
    path = tmp_path / "rates.xlsx"
    make_workbook(path)

    provider = AnnualRateProvider(path, 2025)

    assert provider.get_rate("2025-02-01", "USD") == Decimal("521.59")
    assert provider.get_rate("2025-02-01", "KZT") == Decimal("1")
    assert provider.source_for("USD").rate_cell == "E4"


def test_annual_rate_provider_rejects_missing_year(tmp_path: Path) -> None:
    path = tmp_path / "rates.xlsx"
    make_workbook(path)

    with pytest.raises(AnnualRateError, match="Year 2024"):
        AnnualRateProvider(path, 2024)


def test_annual_rate_provider_rejects_invalid_rate(tmp_path: Path) -> None:
    path = tmp_path / "rates.xlsx"
    make_workbook(path, rate="not-a-rate")

    with pytest.raises(AnnualRateError, match="USD.*invalid"):
        AnnualRateProvider(path, 2025)
