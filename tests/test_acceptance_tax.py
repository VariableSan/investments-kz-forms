from datetime import date
from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from kz_tax_report.report_builder import write_xlsx
from kz_tax_report.tax_engine import RulesError, calculate_report, load_rules


class FakeRates:
    def __init__(self, rates: dict[tuple[str, str], Decimal]) -> None:
        self.rates = rates
        self.calls: list[tuple[str, str]] = []

    def get_rate(self, rate_date: date | str, currency: str) -> Decimal:
        key = (str(rate_date), currency)
        self.calls.append(key)
        return self.rates[key]


def rules_file(tmp_path: Path, *, approved: bool = False) -> Path:
    path = tmp_path / "rules.yaml"
    path.write_text(
        f"""
        year: 2025
        approved: {str(approved).lower()}
        citation: Synthetic citation
        conversion_policy: NBK rate on payment date
        form_labels:
          dividends: Foreign dividends
          realized_gains: Foreign capital gains
          exempt_gains: Exempt Freedom gains
        exemptions: [AIX ETN]
        tax:
          rate: '0.10'
          foreign_tax_credit: true
        income:
          dividends_line: 'B.01.4'
          realized_gains_line: 'A.01.02'
          exempt_gains_line: 'E.adjustment'
        """,
        encoding="utf-8",
    )
    return path


def sections() -> dict[str, pd.DataFrame]:
    return {
        "Дивиденды": pd.DataFrame(
            [
                {
                    "currency": "USD",
                    "date": "2024-12-31",
                    "amount": "9",
                    "source_file": "a.csv",
                    "source_row": 2,
                    "description": "old",
                },
                {
                    "currency": "USD",
                    "date": "2025-02-01",
                    "amount": "100",
                    "source_file": "a.csv",
                    "source_row": 3,
                    "description": "current",
                },
            ]
        ),
        "Удерживаемый налог": pd.DataFrame(
            [
                {
                    "currency": "USD",
                    "date": "2025-02-01",
                    "amount": "15",
                    "source_file": "a.csv",
                    "source_row": 4,
                    "description": "tax",
                }
            ]
        ),
        "Реализованная и нереализованная П/У: отчет об эффективности": pd.DataFrame(
            columns=[
                "asset_class",
                "symbol",
                "realized_total",
                "source_file",
                "source_row",
            ]
        ),
    }


def test_dated_fx_conversion_and_year_filter(tmp_path: Path) -> None:
    rates = FakeRates({("2025-02-01", "USD"): Decimal("500")})
    report = calculate_report(
        year=2025,
        rules=load_rules(rules_file(tmp_path), require_approved=False),
        ibkr_sections=sections(),
        freedom_transactions=pd.DataFrame(columns=["transaction_type", "profit"]),
        f1042s_records=pd.DataFrame(
            columns=["tax_year", "income_code", "gross_income", "federal_tax_withheld"]
        ),
        rate_provider=rates,
    )

    assert report.taxable_dividends == Decimal("50000.00")
    assert report.foreign_tax_credit == Decimal("0.00")
    assert rates.calls == [("2025-02-01", "USD")]
    dividend = next(value for value in report.values if value.category == "dividend")
    assert dividend.foreign_amount == Decimal("100.00")
    assert dividend.kzt_amount == Decimal("50000.00")
    assert dividend.rate_date == "2025-02-01"


def test_1042s_year_filter_keeps_parser_year_records(tmp_path: Path) -> None:
    report = calculate_report(
        year=2025,
        rules=load_rules(rules_file(tmp_path), require_approved=False),
        ibkr_sections=sections(),
        freedom_transactions=pd.DataFrame(columns=["transaction_type", "profit"]),
        f1042s_records=pd.DataFrame(
            [
                {
                    "tax_year": 2025,
                    "income_code": "06",
                    "gross_income": Decimal("100"),
                    "federal_tax_withheld": Decimal("15"),
                    "source_file": "1042-s.pdf",
                    "source_row": 3,
                }
            ]
        ),
        rate_provider=FakeRates({("2025-02-01", "USD"): Decimal("500")}),
    )

    assert report.foreign_tax_credit == Decimal("15")
    assert not any("1042-S was not provided" in warning for warning in report.warnings)


def test_missing_sale_date_fails_instead_of_guessing(tmp_path: Path) -> None:
    freedom = pd.DataFrame(
        [
            {
                "transaction_type": "Продажа",
                "profit": Decimal("10"),
                "source_file": "f.pdf",
                "source_row": 2,
            }
        ]
    )
    with pytest.raises(RulesError, match="Missing conversion date"):
        calculate_report(
            year=2025,
            rules=load_rules(rules_file(tmp_path), require_approved=False),
            ibkr_sections=sections(),
            freedom_transactions=freedom,
            f1042s_records=pd.DataFrame(
                columns=["income_code", "gross_income", "federal_tax_withheld"]
            ),
            rate_provider=FakeRates({("2025-02-01", "USD"): Decimal("500")}),
        )


def test_report_has_paste_sheet_and_final_status(tmp_path: Path) -> None:
    report = calculate_report(
        year=2025,
        rules=load_rules(rules_file(tmp_path), require_approved=False),
        ibkr_sections=sections(),
        freedom_transactions=pd.DataFrame(columns=["transaction_type", "profit"]),
        f1042s_records=pd.DataFrame(
            columns=["income_code", "gross_income", "federal_tax_withheld"]
        ),
    )
    output = tmp_path / "report.xlsx"
    write_xlsx(report, output)
    workbook = load_workbook(output, read_only=True)

    assert "Copy into Form 270.01" in workbook.sheetnames
    assert workbook["Summary"]["B2"].value == "FINAL"
    assert not any("must not be filed" in warning for warning in report.warnings)
