"""Parse annual average exchange rates supplied as an official XLSX."""

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class AnnualRateError(ValueError):
    """Raised when an annual-rate workbook cannot provide a requested rate."""


@dataclass(frozen=True)
class AnnualRateSource:
    """The workbook location that supports an annual exchange rate."""

    file: str
    sheet: str
    currency_cell: str
    rate_cell: str


@dataclass(frozen=True)
class AnnualRate:
    """One validated annual exchange rate and its source location."""

    year: int
    currency: str
    value: Decimal
    source: AnnualRateSource


class AnnualRateProvider:
    """Provide one annual rate per currency from an uploaded workbook."""

    annual = True

    def __init__(self, path: str | Path, year: int) -> None:
        self.path = Path(path)
        self.year = year
        self._rates: dict[str, AnnualRate] = {}
        self._load()

    def get_rate(self, rate_date: date | str, currency: str) -> Decimal:
        """Return the annual rate, ignoring the transaction date by design."""

        del rate_date
        normalized_currency = currency.strip().upper()
        if normalized_currency == "KZT":
            return Decimal("1")
        try:
            return self._rates[normalized_currency].value
        except KeyError as error:
            raise AnnualRateError(
                f"Annual rate for {normalized_currency} is missing in {self.path.name}"
            ) from error

    def source_for(self, currency: str) -> AnnualRateSource | None:
        """Return workbook provenance for a loaded currency rate."""

        rate = self._rates.get(currency.strip().upper())
        return rate.source if rate else None

    def _load(self) -> None:
        if not self.path.is_file():
            raise AnnualRateError(f"Annual rate workbook is missing: {self.path}")

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            candidates = self._find_candidates(workbook)
            if not candidates:
                raise AnnualRateError(
                    f"Year {self.year} or its currency table is missing in "
                    f"{self.path.name}"
                )
            for sheet_name, currency_row, currency_column, year_column in candidates:
                currency = self._currency(
                    workbook[sheet_name].cell(currency_row, currency_column).value
                )
                if not currency or currency in self._rates:
                    continue
                cell = workbook[sheet_name].cell(currency_row, year_column)
                value = self._decimal(cell.value, currency, sheet_name, cell.coordinate)
                self._rates[currency] = AnnualRate(
                    year=self.year,
                    currency=currency,
                    value=value,
                    source=AnnualRateSource(
                        file=self.path.name,
                        sheet=sheet_name,
                        currency_cell=workbook[sheet_name]
                        .cell(currency_row, currency_column)
                        .coordinate,
                        rate_cell=cell.coordinate,
                    ),
                )
        finally:
            workbook.close()

    def _find_candidates(self, workbook: Any) -> list[tuple[str, int, int, int]]:
        candidates: list[tuple[str, int, int, int]] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            year_columns = [
                column
                for column in range(1, sheet.max_column + 1)
                if self._matches_year(sheet.cell(3, column).value)
            ]
            if not year_columns:
                continue
            currency_columns = [
                column
                for column in range(1, sheet.max_column + 1)
                if self._is_currency_header(sheet.cell(3, column).value)
            ]
            currency_column = currency_columns[0] if currency_columns else 3
            for row in range(1, sheet.max_row + 1):
                if self._currency(sheet.cell(row, currency_column).value):
                    for year_column in year_columns:
                        candidates.append(
                            (sheet_name, row, currency_column, year_column)
                        )
        return candidates

    def _matches_year(self, value: object) -> bool:
        return str(value).strip() == str(self.year)

    @staticmethod
    def _is_currency_header(value: object) -> bool:
        return str(value).strip().upper() in {"КОД", "CODE", "CURRENCY", "ВАЛЮТА"}

    @staticmethod
    def _currency(value: object) -> str | None:
        if value is None:
            return None
        normalized = str(value).strip().upper()
        return normalized or None

    @staticmethod
    def _decimal(value: object, currency: str, sheet: str, cell: str) -> Decimal:
        if value is None or isinstance(value, bool):
            raise AnnualRateError(
                f"Annual rate for {currency} is empty at {sheet}!{cell}"
            )
        try:
            result = Decimal(str(value).replace(",", "."))
        except (InvalidOperation, ValueError) as error:
            raise AnnualRateError(
                f"Annual rate for {currency} is invalid at {sheet}!{cell}: {value!r}"
            ) from error
        if result <= 0:
            raise AnnualRateError(
                f"Annual rate for {currency} must be positive at {sheet}!{cell}"
            )
        return result
