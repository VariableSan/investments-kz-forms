"""Source-traceable Markdown and XLSX report writers."""

from pathlib import Path

from openpyxl import Workbook

from kz_tax_report.tax_engine import TaxReport


def write_xlsx(report: TaxReport, path: str | Path) -> None:
    """Write summary, paste-ready form values, provenance, and warnings."""

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Form 270.01 input", "Amount", "Source rule line"])
    summary.append(["Status", report.status, "Human approval required"])
    summary_rows = [
        ("Taxable dividends", report.taxable_dividends, report.rules.dividends_line),
        (
            "Taxable realized gains",
            report.taxable_realized_gains,
            report.rules.realized_gains_line,
        ),
        (
            "Exempt realized gains",
            report.exempt_realized_gains,
            report.rules.exempt_gains_line,
        ),
        ("Tax before foreign credit", report.tax_before_credit, "calculation"),
        ("Foreign tax credit", report.foreign_tax_credit, "calculation"),
        ("Tax due", report.tax_due, "calculation"),
    ]
    for row in summary_rows:
        summary.append(row)

    paste = workbook.create_sheet("Copy into Form 270.01")
    paste.append(["Form label", "KZT amount", "Rule line", "Notes"])
    paste_rows = [
        (
            report.rules.form_labels.get("dividends", "Foreign dividends gross"),
            report.taxable_dividends,
            report.rules.dividends_line,
            "Gross income before foreign withholding",
        ),
        (
            report.rules.form_labels.get("realized_gains", "Foreign realized gains"),
            report.taxable_realized_gains,
            report.rules.realized_gains_line,
            "Taxable disposals only",
        ),
        (
            report.rules.form_labels.get("exempt_gains", "Exempt Freedom gains"),
            report.exempt_realized_gains,
            report.rules.exempt_gains_line,
            "Reported and adjusted under configured exemption",
        ),
        (
            "Foreign tax credit",
            report.foreign_tax_credit,
            "calculation",
            "Capped at Kazakhstan tax on corresponding income",
        ),
        (
            "IPN payable",
            report.tax_due,
            "calculation",
            f"{report.status}; verify rules before filing",
        ),
    ]
    for row in paste_rows:
        paste.append(row)

    values = workbook.create_sheet("Values")
    values.append(
        [
            "Category",
            "Amount KZT",
            "Foreign amount",
            "Currency",
            "Annual FX rate",
            "FX source",
            "Source file",
            "Source row",
            "Detail",
        ]
    )
    for value in report.values:
        values.append(
            [
                value.category,
                value.amount,
                value.foreign_amount,
                value.currency,
                value.fx_rate,
                value.fx_source,
                value.source_file,
                value.source_row,
                value.source_detail,
            ]
        )

    assets = workbook.create_sheet("Copy into Form 270.04")
    assets.append(
        [
            "Asset class",
            "Symbol",
            "ISIN",
            "Quantity",
            "Country",
            "Currency",
            "Source file",
            "Source row",
            "Manual completion",
        ]
    )
    for asset in report.assets:
        assets.append(
            [
                asset.get("asset_class", ""),
                asset.get("symbol", ""),
                asset.get("isin", ""),
                asset.get("quantity", ""),
                asset.get("country", ""),
                asset.get("currency", ""),
                asset.get("source_file", ""),
                asset.get("source_row", ""),
                "ISIN and country require manual completion",
            ]
        )

    warnings = workbook.create_sheet("Warnings")
    warnings.append(["Warning"])
    for warning in report.warnings:
        warnings.append([warning])
    workbook.save(path)


def write_markdown(report: TaxReport, path: str | Path) -> None:
    """Write a concise manual-review summary with source references."""

    lines = [
        f"# Form 270.01 inputs for {report.year}",
        "",
        f"STATUS: {report.status}",
        "",
        f"Rules citation: {report.rules.citation}",
        "",
        "## Copy into Form 270.01",
        "",
        "| Form label | KZT amount | Rule line | Notes |",
        "| --- | ---: | --- | --- |",
        f"| {report.rules.form_labels.get('dividends', 'Foreign dividends gross')} | {report.taxable_dividends} | {report.rules.dividends_line} | Gross income before foreign withholding |",
        f"| {report.rules.form_labels.get('realized_gains', 'Foreign realized gains')} | {report.taxable_realized_gains} | {report.rules.realized_gains_line} | Taxable disposals only |",
        f"| {report.rules.form_labels.get('exempt_gains', 'Exempt Freedom gains')} | {report.exempt_realized_gains} | {report.rules.exempt_gains_line} | Reported and adjusted under configured exemption |",
        f"| Foreign tax credit | {report.foreign_tax_credit} | calculation | Capped at Kazakhstan tax on corresponding income |",
        f"| IPN payable | {report.tax_due} | calculation | {report.status}; verify rules before filing |",
        "",
        "| Input | Amount | Rule line |",
        "| --- | ---: | --- |",
        f"| Taxable dividends | {report.taxable_dividends} | {report.rules.dividends_line} |",
        f"| Taxable realized gains | {report.taxable_realized_gains} | {report.rules.realized_gains_line} |",
        f"| Exempt realized gains | {report.exempt_realized_gains} | {report.rules.exempt_gains_line} |",
        f"| Tax before foreign credit | {report.tax_before_credit} | calculation |",
        f"| Foreign tax credit | {report.foreign_tax_credit} | calculation |",
        f"| Tax due | {report.tax_due} | calculation |",
        "",
        "## Source values",
        "",
        "| Category | Amount KZT | Foreign amount | Currency | Rate | Source |",
        "| --- | ---: | ---: | --- | ---: | --- |",
    ]
    lines.extend(
        f"| {value.category} | {value.amount} | {value.foreign_amount or ''} | {value.currency} | {value.fx_rate or ''} | {value.source_file}:{value.source_row} |"
        for value in report.values
    )
    lines.extend(
        [
            "",
            "## Copy into Form 270.04",
            "",
            "| Asset class | Symbol | ISIN | Quantity | Country | Source |",
            "| --- | --- | --- | ---: | --- | --- |",
        ]
    )
    lines.extend(
        f"| {asset.get('asset_class', '')} | {asset.get('symbol', '')} | {asset.get('isin', '')} | {asset.get('quantity', '')} | {asset.get('country', '')} | {asset.get('source_file', '')}:{asset.get('source_row', '')} |"
        for asset in report.assets
    )
    lines.extend(["", "## Warnings", ""])
    lines.extend(f"- {warning}" for warning in report.warnings or ("None",))
    Path(path).write_text("\n".join(lines) + "\n", encoding="utf-8")
